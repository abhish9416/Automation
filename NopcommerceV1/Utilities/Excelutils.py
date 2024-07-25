import openpyxl
from openpyxl.styles import PatternFill


def getMaxRow(file_name , Sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    return sheet.max_row

def getMaxColumn(file_name, Sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    return sheet.max_column

def read_Data(file_name,Sheet_name,row_num,cloumn_num):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    return sheet.cell(row_num,cloumn_num).value

def write_Data(file_name,Sheet_name,row_num,column_num,data):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    sheet.cell(row_num,column_num).value = data
    workbook.save(file_name)

def fill_green_color(file_name,Sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    greenfill = PatternFill(start_color='00FF00',end_color='00FF00',patternType='solid')
    sheet.cell(row_num,column_num).fill = greenfill
    workbook.save(file_name)

def fill_red_colour(file_name,Sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    redfill = PatternFill(start_color='FF0000',end_color='FF0000',patternType='solid')
    sheet.cell(row_num,column_num).fill = redfill
    workbook.save(file_name)

def next_empty_row(file_name,Sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[Sheet_name]
    for row in range(1,sheet.max_row+1):
        if sheet.cell(row=row, column=1).value is None:
            return row
    return sheet.max_row+1

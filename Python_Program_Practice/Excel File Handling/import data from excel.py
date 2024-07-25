import openpyxl

workbook = openpyxl.load_workbook("testdata.xlsx")

sheet = workbook.active

# cell = sheet.cell(row=1,column=1)
# print(cell.value)

data = []
for i in sheet.iter_cols(values_only=True):
    data.append(i)

print(data)
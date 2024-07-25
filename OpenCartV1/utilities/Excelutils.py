import openpyxl
from openpyxl.styles import PatternFill


def getrowcount(filename,sheet_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    return sheet.max_row

def getcolumncount(filename,sheet_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    return sheet.max_column

def readData(filename,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num,column_num).value

def writedata(filename,sheet_name,row_num,column_num,data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    sheet.cell(row_num,column_num).value = data
    workbook.save(filename)

def fillgreencolour(file_name,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    greenfill = PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
    sheet.cell(row_num,column_num).fill = greenfill
    workbook.save(file_name)

def fillredcolour(file_name,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    redfill = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    sheet.cell(row_num,column_num).fill = redfill
    workbook.save(file_name)

# def get_next_empty_row(file_name, sheet_name):
#     workbook = openpyxl.load_workbook(file_name)
#     sheet = workbook[sheet_name]
#     for row in range(2, sheet.max_row+1):
#         return sheet.max_row + 1

def get_next_empty_row(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    for row in range(1, sheet.max_row+1):
        if sheet.cell(row=row, column=1).value is None:
            return row
    return sheet.max_row + 1

# rowcount = getrowcount('../testData/Userdetail.xlsx',"Sheet1")
# print(rowcount)
#
# column = getcolumncount("../testData/Userdetail.xlsx","Sheet1")
# print(column)
